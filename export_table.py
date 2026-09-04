"""Bot-side Excel export using the existing lightweight Ubuntu dependency."""
import tempfile
import os
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from config import WEEKDAY_SHORT_RU

def build_xlsx(dates: list[str], rows: list[dict]) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Посещения_bot'
    ws.append(['Ответы бота: Y — «Приду», N — «Не приду».'])
    ws.append(['ID участника', 'Telegram ID', 'ФИО', 'Телеграм', 'Телефон'] +
              [datetime.fromisoformat(d) for d in dates])
    ws.append(['', '', '', '', 'День недели'] + [WEEKDAY_SHORT_RU[datetime.fromisoformat(d).weekday()] for d in dates])
    for row in rows:
        ws.append([row['participant_id'], str(row['telegram_id'] or ''), row['full_name'],
                   '@'+row['username'] if row['username'] else '', row['phone']] +
                  [row['marks'].get(day, '') for day in dates])
        for cell in ws[ws.max_row][:5]:
            cell.data_type = 's'
    for row in ws.iter_rows(min_row=2, max_row=3):
        for cell in row:
            cell.fill = PatternFill('solid', fgColor='243746')
            cell.font = Font(bold=True, color='FFFFFF')
            cell.alignment = Alignment(horizontal='center')
    for cell in ws[2][5:]:
        cell.number_format = 'dd.mm.yyyy'
    for idx, width in enumerate((18, 18, 36, 20, 20), 1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    for idx in range(6, len(dates)+6):
        ws.column_dimensions[get_column_letter(idx)].width = 13
    ws.freeze_panes = 'F4'
    ws.sheet_view.showGridLines = False
    info = wb.create_sheet('Как использовать')
    for line in [
        'Y означает «Приду», N — «Не приду», пусто — ответа ещё нет.',
        'Календарь начинается 1 августа 2026; даты без ответов остаются пустыми.',
        'ID участника постоянный в этой базе; Telegram ID получен от Telegram.',
        'ФИО и @username могут меняться, поэтому объединяйте записи по ID.',
        'Все дальнейшие расчёты можно делать на других листах по четырёхзначному ID.',
        'Для людей из старого Excel один раз сопоставьте Telegram ID в «Участники»; не объединяйте по имени.',
        'Если в один день несколько тренировок, Y означает хотя бы один положительный ответ.',
    ]:
        info.append([line])
    info.column_dimensions['A'].width = 105
    for row in info:
        row[0].alignment = Alignment(wrap_text=True, vertical='top')
        info.row_dimensions[row[0].row].height = 34
    fd, path = tempfile.mkstemp(suffix='.xlsx', prefix='attendance_')
    os.close(fd)
    wb.save(path)
    wb.close()
    return path
