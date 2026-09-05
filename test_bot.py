import asyncio
import os
from pathlib import Path
import tempfile
import unittest
from datetime import datetime, timedelta
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

import db
import events
import utils
from scheduler import tick
from handlers_poll import process_answer
from export_table import build_xlsx
from google_sheet import merge_grid, _active_formula
from openpyxl import load_workbook

class DataTests(unittest.IsolatedAsyncioTestCase):
    async def asyncSetUp(self):
        self.tmp = tempfile.TemporaryDirectory()
        self.old_path = db.DB_PATH
        db.DB_PATH = str(Path(self.tmp.name) / 'test.db')
        await db.init_db()
        self.now = utils.TZ.localize(datetime(2026,9,7,18,0))
        self.clock = patch('utils.now', return_value=self.now)
        self.mock_now = self.clock.start()
        self.pid = await db.register_participant(77,'tester','Test User','+79991234567')
        await db.set_active(self.pid,True)

    async def asyncTearDown(self):
        self.clock.stop()
        await db.close_db()
        db._conn = None
        db.DB_PATH = self.old_path
        self.tmp.cleanup()

    async def slot(self, time='19:00', day='2026-09-07'):
        return await events.save_slot(datetime.fromisoformat(day).weekday(),time,day,starts_on=day)

    async def test_one_poll_per_session_and_restart(self):
        await self.slot()
        bot=SimpleNamespace(send_message=AsyncMock(return_value=SimpleNamespace(message_id=123)))
        await tick(bot)
        await db.close_db()
        await db.init_db()
        await tick(bot)
        self.assertEqual(bot.send_message.await_count,1)

    async def test_disabled_and_unregistered_not_contacted(self):
        await self.slot()
        await db.set_active(self.pid,False)
        await db.create_stub_participant('waiting',None)
        bot=SimpleNamespace(send_message=AsyncMock())
        await tick(bot)
        bot.send_message.assert_not_awaited()

    async def test_retry_after_failed_delivery(self):
        await self.slot()
        bot=SimpleNamespace(send_message=AsyncMock(side_effect=[RuntimeError('offline'),SimpleNamespace(message_id=1)]))
        with self.assertLogs('scheduler',level='ERROR'):
            await tick(bot)
        await tick(bot)
        self.assertEqual(bot.send_message.await_count,2)

    async def test_midnight_and_weekly_start_boundary(self):
        self.mock_now.return_value=utils.TZ.localize(datetime(2026,9,7,23,30))
        await events.save_slot(1,'00:30',starts_on='2026-09-15')
        bot=SimpleNamespace(send_message=AsyncMock(return_value=SimpleNamespace(message_id=1)))
        await tick(bot)
        bot.send_message.assert_not_awaited()
        await self.slot('00:30','2026-09-08')
        await tick(bot)
        bot.send_message.assert_awaited_once()

    async def test_edit_and_cancel_take_effect(self):
        sid=await self.slot()
        await events.save_slot(0,'22:00','2026-09-07',sid)
        bot=SimpleNamespace(send_message=AsyncMock())
        await tick(bot)
        bot.send_message.assert_not_awaited()
        await events.delete_slot(sid)
        self.assertIsNone(await events.get_slot(sid))

    async def test_multiple_sessions_one_day(self):
        a=await self.slot('18:30')
        b=await self.slot('19:00')
        bot=SimpleNamespace(send_message=AsyncMock(return_value=SimpleNamespace(message_id=1)))
        await tick(bot)
        self.assertEqual(bot.send_message.await_count,2)
        first=await events.response_for(self.pid,a,self.now+timedelta(minutes=30))
        second=await events.response_for(self.pid,b,self.now+timedelta(hours=1))
        await db._c().execute("UPDATE responses SET status='no' WHERE id=?",(first['id'],))
        await db._c().execute("UPDATE responses SET status='yes' WHERE id=?",(second['id'],))
        await db._c().commit()
        _,rows=await events.summary()
        self.assertEqual(rows[0]['marks']['2026-09-07'],'Y')

    async def test_calendar_runs_from_august_through_december(self):
        dates, _ = await events.summary()
        self.assertEqual(dates[0], '2026-08-01')
        self.assertEqual(dates[-1], '2026-12-31')
        self.assertEqual(len(dates), 153)

    async def test_duplicate_slot_rejected(self):
        await self.slot()
        with self.assertRaises(ValueError):
            await self.slot()

    async def test_answer_ownership_change_and_closed_poll(self):
        sid=await self.slot()
        start=self.now+timedelta(hours=1)
        r=await events.response_for(self.pid,sid,start)
        callback=SimpleNamespace(id='query',data=f"r:{r['id']}:yes",answer=AsyncMock(),
            from_user=SimpleNamespace(id=999),
            message=SimpleNamespace(answer=AsyncMock(),edit_text=AsyncMock()))
        await process_answer(callback)
        row=await (await db._c().execute('SELECT status FROM responses')).fetchone()
        self.assertEqual(row['status'],'pending')
        callback.from_user.id=77
        await process_answer(callback)
        dates,rows=await events.summary()
        self.assertEqual(rows[0]['marks']['2026-09-07'],'Y')
        callback.data=f"r:{r['id']}:no"
        with patch('handlers_poll.google_sheet.queue') as queue:
            await process_answer(callback)
            queue.assert_called_once_with()
        dates,rows=await events.summary()
        self.assertEqual(rows[0]['marks']['2026-09-07'],'N')
        await events.delete_slot(sid)
        callback.data=f"r:{r['id']}:yes"
        await process_answer(callback)
        dates,rows=await events.summary()
        self.assertEqual(rows[0]['marks']['2026-09-07'],'N')
        callback.message.answer.assert_awaited_once()

    async def test_four_digit_ids_are_unique_and_survive_restart(self):
        second = await db.register_participant(88,'second','Second User','+79990000000')
        first_code = (await db.get_participant(self.pid))['public_id']
        second_code = (await db.get_participant(second))['public_id']
        self.assertRegex(str(first_code), r'^\d{4}$')
        self.assertRegex(str(second_code), r'^\d{4}$')
        self.assertNotEqual(first_code, second_code)
        await db.close_db()
        await db.init_db()
        self.assertEqual((await db.get_participant(self.pid))['public_id'], first_code)
        self.assertEqual((await db.get_participant(second))['public_id'], second_code)
        self.assertTrue(await db.adopt_public_id(second, 8001))
        self.assertFalse(await db.adopt_public_id(self.pid, 8001))
        self.assertEqual((await db.get_participant(second))['public_id'], 8001)

    async def test_invited_registration_preserves_disabled_flag(self):
        stub=await db.create_stub_participant('someone',None)
        await db.set_active(stub,False)
        await db.register_participant(88,'someone','Name','1234567',stub)
        p=await db.get_participant(stub)
        self.assertFalse(p['is_active'])
        new=await db.register_participant(99,None,'New','1234567')
        self.assertFalse((await db.get_participant(new))['is_active'])

    async def test_legacy_history_and_export_shape(self):
        aid=await db.get_or_create_attendance(self.pid,'2026-09-01')
        await db.update_attendance_status(aid,'yes')
        await db.set_active(self.pid,False)
        dates,rows=await events.summary()
        rows[0]['full_name']='=UNTRUSTED()'
        path=build_xlsx(dates,rows)
        try:
            wb=load_workbook(path)
            ws=wb.active
            self.assertEqual(wb.sheetnames,['Посещения_bot','Как использовать'])
            self.assertEqual([c.value for c in ws[2]][:6],
                             ['ID участника','Telegram ID','ФИО','Телеграм','Телефон','flag_active'])
            self.assertEqual(ws.cell(4,7+dates.index('2026-09-01')).value,'Y')
            self.assertEqual(ws['C4'].data_type,'s')
            self.assertEqual(ws['B4'].value,'77')
            self.assertEqual(ws['A4'].value,'1001')
            self.assertTrue(ws['F4'].value.startswith('=IF(COUNTIFS('))
            self.assertEqual(ws['G2'].value,datetime(2026,8,1))
            self.assertEqual(ws['G3'].value,'Сб')
            self.assertEqual(ws.freeze_panes,'G4')
            wb.close()
        finally:
            os.unlink(path)


class GoogleSheetMergeTests(unittest.TestCase):
    def test_active_formula_uses_rolling_thirty_days_and_russian_separators(self):
        formula = _active_formula(3, 'FC')
        self.assertIn('TODAY()-30', formula)
        self.assertIn('G3:FC3', formula)
        self.assertIn(';"Y"', formula)
        self.assertNotIn(',"Y"', formula)

    def test_historical_row_is_reused_by_telegram_id(self):
        existing = [
            ['ID участника','Telegram ID','ФИО','Телеграм','Телефон','01.08.2026'],
            ['','','','','День недели','Сб'],
            ['8001','77','Старое имя','','','Y'],
        ]
        rows = [{
            'internal_id': 1, 'participant_id': '1001', 'telegram_id': 77,
            'full_name': 'Новое имя', 'username': 'newname', 'phone': '+79991234567',
            'marks': {'2026-08-01': 'Y', '2026-09-07': 'N'},
        }]
        grid, adoptions = merge_grid(existing, ['2026-08-01','2026-09-07'], rows)
        self.assertEqual(adoptions, [(1,8001)])
        self.assertEqual(grid[2][:5], ['8001','77','Новое имя','@newname','+79991234567'])
        self.assertEqual(grid[2][5:], ['', 'Y','N'])
        self.assertEqual(grid[0][5], 'flag_active')
        self.assertEqual(grid[0][-1], '07.09.2026')

    def test_new_participant_appends_without_touching_history(self):
        existing = [
            ['ID участника','Telegram ID','ФИО','Телеграм','Телефон','01.08.2026'],
            ['','','','','День недели','Сб'],
            ['8001','','История','','','Y'],
        ]
        rows = [{
            'internal_id': 2, 'participant_id': '1001', 'telegram_id': 88,
            'full_name': '=Не формула', 'username': '', 'phone': '1234567',
            'marks': {'2026-08-01': ''},
        }]
        grid, adoptions = merge_grid(existing, ['2026-08-01'], rows)
        self.assertEqual(adoptions, [])
        self.assertEqual(grid[2][0:7], ['8001','','История','','','','Y'])
        self.assertEqual(grid[3][0:5], ['1001','88','=Не формула','','1234567'])

    def test_existing_flag_column_is_not_treated_as_a_date(self):
        existing = [
            ['ID участника','Telegram ID','ФИО','Телеграм','Телефон','flag_active','01.08.2026'],
            ['','','','','День недели','За последние 30 дней','Сб'],
            ['8001','','История','','','active','Y'],
        ]
        grid, _ = merge_grid(existing, ['2026-08-01'], [])
        self.assertEqual(grid[0][5:7], ['flag_active','01.08.2026'])
        self.assertEqual(grid[2][5:7], ['', 'Y'])

    def test_changed_or_duplicate_keys_stop_sync(self):
        with self.assertRaises(ValueError):
            merge_grid([['ID','Telegram ID','ФИО','Телеграм','Телефон']], [], [])
        duplicate = [
            ['ID участника','Telegram ID','ФИО','Телеграм','Телефон'],
            ['','','','','День недели'],
            ['8001','','A','',''], ['8001','','B','',''],
        ]
        with self.assertRaises(ValueError):
            merge_grid(duplicate, [], [])

class RoutingTest(unittest.IsolatedAsyncioTestCase):
    async def test_registration_calendar_and_role_boundaries(self):
        from aiogram import Bot, Dispatcher
        from aiogram.client.session.base import BaseSession
        from aiogram.fsm.storage.memory import SimpleEventIsolation
        from aiogram.types import Message, Update
        from config import TRAINER_IDS
        import handlers_menu, handlers_trainer, handlers_poll, handlers_registration, handlers_manual
        import interaction, background
        old_trainers=set(TRAINER_IDS)
        TRAINER_IDS.clear()
        TRAINER_IDS.add(1)
        tmp=tempfile.TemporaryDirectory()
        old_path=db.DB_PATH
        db.DB_PATH=str(Path(tmp.name)/'routing.db')
        await db.init_db()
        sent=[]
        latest_ui={}
        class Session(BaseSession):
            async def close(self): pass
            async def make_request(self, bot, method, timeout=None):
                sent.append(method)
                if method.__api_method__ in {'sendMessage','editMessageText'}:
                    mid=getattr(method,'message_id',None) or len(sent)
                    if getattr(method,'reply_markup',None) and hasattr(method.reply_markup,'inline_keyboard'):
                        latest_ui[method.chat_id]=mid
                    return Message(message_id=mid,date=datetime.now(),
                        chat={'id':method.chat_id,'type':'private'},text=method.text)
                return True
            async def stream_content(self,*args,**kwargs):
                yield b''
        bot=Bot('123456:TEST_TOKEN',session=Session())
        dp=Dispatcher(events_isolation=SimpleEventIsolation())
        interaction.install(dp)
        from aiogram import F
        dp.message.filter(F.chat.type=='private')
        dp.callback_query.filter(F.message.chat.type=='private')
        for r in (handlers_menu.router,handlers_trainer.router,handlers_poll.router,handlers_registration.router,handlers_menu.fallback):
            dp.include_router(r)
        seq=0
        async def feed(text=None,callback=None,uid=1,chat_type='private',message_id=None):
            nonlocal seq
            seq+=1
            user={'id':uid,'is_bot':False,'first_name':'Test','username':'testuser'}
            msg={'message_id':seq,'date':int(datetime.now().timestamp()),
                 'chat':{'id':uid,'type':chat_type},'from':user,'text':text or 'button'}
            if text and text.startswith('/'):
                msg['entities']=[{'type':'bot_command','offset':0,'length':len(text.split()[0])}]
            data={'update_id':seq}
            if callback:
                msg['message_id']=message_id or latest_ui.get(uid,seq)
                data['callback_query']={'id':str(seq),'from':user,'chat_instance':'test','data':callback,'message':msg}
            else:
                data['message']=msg
            await dp.feed_update(bot,Update.model_validate(data))
        try:
            await feed('/start')
            await feed('Test Full Name')
            self.assertIsNone(await db.get_participant_by_telegram_id(1))
            await feed('Иванов Иван Иванович')
            await feed('+79991234567')
            self.assertIsNotNone(await db.get_participant_by_telegram_id(1))
            await feed('0')
            self.assertIn('Выберите',sent[-1].text)
            await feed('➕ Тренировка')
            day=utils.today()+timedelta(days=2)
            await feed(callback=f'date:{day.isoformat()}')
            await feed(callback='hour:19')
            await feed(callback='time:19:30')
            await feed(callback='repeat:weekly')
            await feed(callback='save_training')
            slots=await db.list_schedule()
            self.assertEqual(len(slots),1)
            self.assertEqual(slots[0]['starts_on'],day.isoformat())
            await feed(callback='save_training')
            self.assertEqual(len(await db.list_schedule()),1)
            await feed(callback=f"delete_confirm:{slots[0]['id']}",uid=2)
            self.assertIsNotNone(await events.get_slot(slots[0]['id']))
            before=len(sent)
            await feed('/table',chat_type='group')
            self.assertEqual(len(sent),before)
            with patch('handlers_trainer.google_sheet.configured',return_value=True), \
                 patch('handlers_trainer.background.start') as export_start:
                await feed('/table')
                export_start.assert_called_once()
                self.assertIn('docs.google.com/spreadsheets',sent[-1].text)
            await feed('➕ Участник')
            await feed('@someone @tester 54545234')
            self.assertIsNotNone(await db.get_participant_by_username('someone'))
            p=await db.get_participant_by_telegram_id(54545234)
            self.assertFalse(p['is_registered'])
            self.assertTrue(p['is_active'])
            await feed('/start',uid=54545234)
            await feed('Петров Пётр Петрович',uid=54545234)
            await feed('+79991234567',uid=54545234)
            self.assertTrue((await db.get_participant_by_telegram_id(54545234))['is_registered'])
            await feed('/id',uid=54545234)
            self.assertIn('54545234',sent[-1].text)
            # Old wizard buttons cannot edit a newly started wizard.
            await feed('/training')
            old_ui=latest_ui[1]
            await feed('/cancel')
            await feed('/training')
            await feed(callback=f'date:{day.isoformat()}',message_id=old_ui)
            self.assertIn('предыдущего',next(m.text for m in reversed(sent) if hasattr(m,'text')))
            await feed(callback=f'date:{day.isoformat()}')
            # A global action must work while waiting for a time value.
            await feed('👥 Участники')
            self.assertIn('Участники',next(m.text for m in reversed(sent) if hasattr(m,'text')))
            await feed('/poll_now')
            await feed(callback='manual_select:0')
            with patch('handlers_manual.kick') as kick_mock:
                await feed(callback='manual_confirm')
                kick_mock.assert_called_once()
                await feed(callback='manual_confirm')
                kick_mock.assert_called_once()
            request=await (await db._c().execute('SELECT * FROM manual_polls')).fetchone()
            self.assertEqual(request['status'],'pending')
            await feed('/training')
            await feed(callback=f'date:{day.isoformat()}')
            await feed('/start')
            self.assertIn('С возвращением',next(m.text for m in reversed(sent) if hasattr(m,'text')))
            await feed(callback=f"delete_slot:{slots[0]['id']}")
            deletion_ui=latest_ui[1]
            await feed('/cancel')
            await feed(callback=f"delete_confirm:{slots[0]['id']}",message_id=deletion_ui)
            self.assertIsNotNone(await events.get_slot(slots[0]['id']))
            await feed(callback=f"delete_slot:{slots[0]['id']}")
            await feed(callback=f"delete_confirm:{slots[0]['id']}")
            self.assertIsNone(await events.get_slot(slots[0]['id']))
        finally:
            await background.close()
            await dp.storage.close()
            await bot.session.close()
            await db.close_db()
            db._conn=None
            db.DB_PATH=old_path
            TRAINER_IDS.clear()
            TRAINER_IDS.update(old_trainers)
            tmp.cleanup()

if __name__=='__main__':
    unittest.main()
