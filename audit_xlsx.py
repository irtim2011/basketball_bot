"""Audit a complete Google workbook export without requiring Excel."""
import json
import re
import sys
from zipfile import ZipFile
from openpyxl import load_workbook

REQUIRED = {'Посещения_bot', 'Посещения', 'Тарифы', 'Покупки тарифов',
            'Номинальная доходность', 'Фактическая прибыль', 'Аналитика_тех',
            'Справочник_клиентов', 'RUN'}
SHEET_REF = re.compile(r"(?:'((?:[^']|'')+)'|([\w.]+))!")


def audit(path):
    errors = []
    with ZipFile(path) as archive:
        external = [name for name in archive.namelist() if name.startswith('xl/externalLinks/')]
    if external: errors.append('externalLinks present')
    formulas = load_workbook(path, read_only=True, data_only=False)
    values = load_workbook(path, read_only=True, data_only=True)
    names = set(formulas.sheetnames)
    if REQUIRED - names: errors.append('Missing sheets: ' + ', '.join(sorted(REQUIRED - names)))
    count = 0
    for sheet in formulas:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'e': errors.append(f'{sheet.title}!{cell.coordinate}: {cell.value}')
                if cell.data_type != 'f': continue
                count += 1
                formula = str(cell.value)
                if '#REF!' in formula or re.search(r'\[[^\]]+\]|IMPORTRANGE|https?://', formula, re.I):
                    errors.append(f'{sheet.title}!{cell.coordinate}: external or broken formula')
                for quoted, plain in SHEET_REF.findall(formula):
                    target = (quoted or plain).replace("''", "'")
                    if target not in names: errors.append(f'{sheet.title}!{cell.coordinate}: missing sheet {target}')
    for sheet in values:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.data_type == 'e': errors.append(f'{sheet.title}!{cell.coordinate}: cached {cell.value}')
    formulas.close(); values.close()
    return {'sheets': sorted(names), 'formula_count': count, 'externalLinks': external,
            'errors': sorted(set(errors)), 'passed': not errors}


if __name__ == '__main__':
    result = audit(sys.argv[1]); print(json.dumps(result, ensure_ascii=False, indent=2))
    sys.exit(0 if result['passed'] else 1)
